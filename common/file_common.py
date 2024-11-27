import json
import os
import re
import openpyxl as op
from datetime import datetime
from PyQt5 import QtWidgets


def check_exist_file(file_path):
    if not os.path.exists(file_path):
        print(file_path, 'is not found')
        return False
    else:
        return True


def load_json(file_path):
    if check_exist_file(file_path):
        if re.search('\.json', file_path) is not None:
            with open(file_path, 'r', encoding='cp932', errors='ignore') as json_file:
                data = json.load(json_file)
            return data
        else:
            print('Input file must be a .json file')


def load_excel(file_path):
    if check_exist_file(file_path):
        if re.search('\.xlsx', file_path) is not None:
            wb = op.load_workbook(file_path)
            ws = wb.active
            import_data = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    value = cell.internal_value
                    row_data.append(value if value is not None else '')
                import_data.append(row_data)
            return import_data


def open_dialog_file(file_type="Excel (*.xlsx );;All Files (*)"):
    QFileDialog = QtWidgets.QFileDialog()
    options = QFileDialog.Options()
    options |= QFileDialog.DontUseNativeDialog
    fileName, _ = QFileDialog.getOpenFileName(QFileDialog, "Select File", os.path.expanduser('.'),
                                              file_type, options=options)
    if fileName:
        print(fileName)
        return fileName
    else:
        return None


def open_dialog_folder():
    QFileDialog = QtWidgets.QFileDialog()
    options = QFileDialog.Options()
    options |= QFileDialog.DontUseNativeDialog
    dir = QFileDialog.getExistingDirectory(QFileDialog, "Select Directory")
    if dir:
        print(dir)
        return dir
    else:
        return None


def save_excel(table, column_name, data, destination=''):
    wb = op.Workbook()
    ws = wb.active
    ws.append(column_name)
    for item in data:
        ws.append(item)
    if destination != '':
        if not check_exist_file(destination):
            os.makedirs(destination)

    time = datetime.now()
    file_name = destination + f'{table}'.capitalize() + "_{}_{}_{}_{}_{}_{}.xlsx". \
        format(time.day, time.month, time.year, time.hour, time.minute, time.second)
    wb.save(file_name)
    return file_name
